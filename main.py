import yfinance as yf
import pandas as pd
import numpy as np
from ta.trend import MACD
from ta.momentum import ROCIndicator, StochasticOscillator
import openpyxl
import os
import subprocess
from datetime import datetime, timedelta
import sys
import time
import random
import warnings

# Suprimir el warning de urllib3
warnings.filterwarnings('ignore', category=Warning, module='urllib3')

# Configuración
TICKERS = [
    'SPY', 'TLT', 'QQQ', 'SLV', 'GLD', 'USO', 'XLE', 'XLRE', 'XLI', 
    'XLF', 'XLB', 'XLY', 'XLK', 'XLP', 'XLV', 'XLU', 'UUP', 'MTUM', 
    'MOAT', 'SPYV', 'SPYG', 'RSP', 'IWO', 'IWN', 'GMF', 'IBIT', 'FXI'
]

# Constantes
ROC_WINDOW = 26
MACD_FAST = 12
MACD_SLOW = 26
MACD_SIGNAL = 9
# Nuevas constantes para MACD trimestral
MACD_TRI_FAST = 36
MACD_TRI_SLOW = 78
MACD_TRI_SIGNAL = 21
STOCH_WINDOW = 89
STOCH_SMOOTH = 3 
EXCEL_FILE = 'resultados.xlsx'

# Configuración de reintentos y tiempos de espera
MAX_RETRIES = 3  # Mínimo de reintentos
BASE_DELAY = 1   # Mínimo tiempo de espera
MAX_DELAY = 5    # Mínimo tiempo máximo
BATCH_SIZE = 27  # Procesar todos los tickers de una vez
INTER_BATCH_DELAY = (0, 0)  # Sin pausa entre lotes
INTER_DOWNLOAD_DELAY = (0, 0)  # Sin pausa entre descargas

# Fechas para descarga de datos
END_DATE = datetime.now()
START_DATE = END_DATE - timedelta(days=3650)  # 10 años

# Agregar al inicio del archivo, después de los imports
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def print_header():
    """Imprime un encabezado atractivo para la aplicación"""
    print("\n" + "=" * 60)
    print(f"{Colors.HEADER}{Colors.BOLD}ANÁLISIS TÉCNICO DE MERCADOS{Colors.ENDC}".center(60))
    print("=" * 60)
    print(f"{Colors.BLUE}Fecha: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}{Colors.ENDC}")
    print(f"{Colors.BLUE}Tickers analizados: {len(TICKERS)}{Colors.ENDC}")
    print(f"{Colors.BLUE}Período: {START_DATE.strftime('%d-%m-%Y')} - {END_DATE.strftime('%d-%m-%Y')}{Colors.ENDC}")
    print("=" * 60 + "\n")

def download_data_with_retry(ticker, period):
    """Descarga datos históricos con sistema de reintentos"""
    for attempt in range(MAX_RETRIES):
        try:
            data = yf.download(
                ticker, 
                start=START_DATE, 
                end=END_DATE, 
                interval=period, 
                auto_adjust=True,
                progress=False,
                threads=False
            )
            
            if not data.empty:
                return data
            else:
                if ticker == 'IBIT':
                    print(f"⚠️ No hay datos disponibles para {ticker}")
                return pd.DataFrame()
            
        except Exception as e:
            if attempt == MAX_RETRIES - 1:
                if ticker == 'IBIT':
                    print(f"❌ Error descargando {ticker}: {str(e)}")
                return pd.DataFrame()
    
    return pd.DataFrame()

def download_data(ticker, period):
    """Función wrapper para la descarga de datos"""
    try:
        return download_data_with_retry(ticker, period)
    except Exception as e:
        print(f"Error descargando datos para {ticker}: {str(e)}")
        return pd.DataFrame()

def calculate_indicators(df):
    """Calcula los indicadores técnicos para el análisis"""
    # ROC (Rate of Change) - Usando el último precio vs el precio de hace 26 períodos
    last_price = df['Close'].iloc[-1].item()
    prev_price = df['Close'].iloc[-ROC_WINDOW-1].item()
    roc = ((last_price - prev_price) / prev_price) * 100
    df['ROC'] = roc
    
    # MACD (Moving Average Convergence Divergence)
    exp1 = df['Close'].ewm(span=MACD_FAST, adjust=False).mean()
    exp2 = df['Close'].ewm(span=MACD_SLOW, adjust=False).mean()
    df['MACD'] = exp1 - exp2
    df['MACD_Signal'] = df['MACD'].ewm(span=MACD_SIGNAL, adjust=False).mean()
    df['MACD_Hist'] = df['MACD'] - df['MACD_Signal']
    
    # Stochastic Oscillator
    low_min = df['Low'].rolling(window=STOCH_WINDOW).min()
    high_max = df['High'].rolling(window=STOCH_WINDOW).max()
    
    # %K (Línea rápida)
    df['Stochastic_K'] = 100 * ((df['Close'] - low_min) / (high_max - low_min))
    
    # %D (Línea lenta - media móvil simple de %K)
    df['Stochastic_D'] = df['Stochastic_K'].rolling(window=STOCH_SMOOTH).mean()
    
    return df

def calculate_trimestral_macd(df):
    """Calcula el MACD con parámetros trimestrales (36,78,21)"""
    # Calcular EMAs usando datos mensuales
    ema_36 = df['Close'].ewm(span=36, adjust=False, min_periods=0).mean()
    ema_78 = df['Close'].ewm(span=78, adjust=False, min_periods=0).mean()
    
    # Guardar las EMAs para comparación
    df['EMA_36'] = ema_36
    df['EMA_78'] = ema_78
    
    # Calcular MACD y señal
    df['MACD_TRI'] = ema_36 - ema_78
    df['MACD_TRI_Signal'] = df['MACD_TRI'].ewm(span=21, adjust=False, min_periods=0).mean()
    df['MACD_TRI_Hist'] = df['MACD_TRI'] - df['MACD_TRI_Signal']
    
    return df

def get_trimestral_signal(df):
    """Determina si el MACD trimestral está en verde o rosa"""
    try:
        macd = df['MACD_TRI'].iloc[-1].item()
        signal = df['MACD_TRI_Signal'].iloc[-1].item()
        return 'verde' if macd > signal else 'rosa'
    except Exception as e:
        return 'rosa'

def calculate_cross_macd(df):
    """Calcula las EMAs para la señal de cruce (12 y 9)"""
    # Calcular EMA de 12 períodos
    df['EMA_12'] = df['Close'].ewm(span=12, adjust=False).mean()
    # Calcular EMA de 9 períodos (Señal)
    df['EMA_9'] = df['Close'].ewm(span=9, adjust=False).mean()
    return df

def get_cross_signal(df):
    """Determina si hay un cruce de EMA 12 sobre EMA 9"""
    try:
        ema12_last = df['EMA_12'].iloc[-1].item()
        ema12_prev = df['EMA_12'].iloc[-2].item()
        ema9_last = df['EMA_9'].iloc[-1].item()
        ema9_prev = df['EMA_9'].iloc[-2].item()
        
        if ema12_prev < ema9_prev and ema12_last > ema9_last:
            return 'azul'
        elif ema12_prev > ema9_prev and ema12_last < ema9_last:
            return 'naranja'
        return None
    except Exception as e:
        return None

def get_macd_signal(df):
    """Determina si MACD está cortado al alza o a la baja"""
    return 'alza' if df['MACD_Hist'].iloc[-1].item() > 0 else 'baja'

def get_stoch_signal(df):
    """Determina las condiciones del Estocástico"""
    try:
        current_k = df['Stochastic_K'].iloc[-1].item()
        current_d = df['Stochastic_D'].iloc[-1].item()
        
        # Verificar si SK>%D
        stoch_condition_up = current_k > current_d
        # Verificar si SK<%D
        stoch_condition_down = current_k < current_d
        
        return {
            'up': stoch_condition_up,
            'down': stoch_condition_down,
            'current_k': current_k
        }
    except Exception as e:
        print(f"Error al calcular señal estocástica: {str(e)}")
        return {'up': False, 'down': False, 'current_k': 0}

def get_resource_path(relative_path):
    """Obtiene la ruta absoluta para recursos empaquetados"""
    try:
        # PyInstaller crea un directorio temporal y almacena la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

def export_to_excel(df):
    """Exporta los resultados a Excel con formato de color"""
    # Obtener la ruta del directorio del ejecutable
    if getattr(sys, 'frozen', False):
        # Si es un ejecutable
        application_path = os.path.dirname(sys.executable)
    else:
        # Si es script Python
        application_path = os.path.dirname(os.path.abspath(__file__))
    
    # Definir la ruta completa del archivo Excel
    excel_path = os.path.join(application_path, EXCEL_FILE)
    
    # Verificar si el archivo ya existe
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws = wb.create_sheet('Sheet')
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
    
    # Estilo para encabezados
    header_style = openpyxl.styles.PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    font_style = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Escribir los encabezados
    headers = ['Ticker', 'ROC', 'Trimestral', 'Mensual', 'Semanal', 'Señal']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_style
        cell.font = font_style
    
    # Definir los colores para los indicadores
    verde = openpyxl.styles.PatternFill(
        start_color="90EE90", end_color="90EE90", fill_type="solid"
    )  # Verde claro
    amarillo = openpyxl.styles.PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )  # Amarillo
    rosa = openpyxl.styles.PatternFill(
        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
    )  # Rosa claro
    azul = openpyxl.styles.PatternFill(
        start_color="87CEEB", end_color="87CEEB", fill_type="solid"
    )  # Azul claro
    naranja = openpyxl.styles.PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )  # Naranja
    
    # Escribir los datos y aplicar colores
    for row, (_, data) in enumerate(df.iterrows(), 2):
        # Ticker
        ws.cell(row=row, column=1, value=data['Ticker'])
        
        # ROC
        roc_cell = ws.cell(row=row, column=2, value=data['ROC'])
        if data['ROC'] > 0:
            roc_cell.font = openpyxl.styles.Font(color="006100")  # Verde oscuro
        else:
            roc_cell.font = openpyxl.styles.Font(color="9C0006")  # Rojo oscuro
        
        # Trimestral (aplicar color)
        cell_trimestral = ws.cell(row=row, column=3)
        if data['Trimestral'] == 'verde':
            cell_trimestral.fill = verde
        else:
            cell_trimestral.fill = rosa
        
        # Mensual (aplicar color)
        cell_mensual = ws.cell(row=row, column=4)
        if data['Mensual'] == 'verde':
            cell_mensual.fill = verde
        elif data['Mensual'] == 'amarillo':
            cell_mensual.fill = amarillo
        else:
            cell_mensual.fill = rosa
            
        # Semanal (aplicar color)
        cell_semanal = ws.cell(row=row, column=5)
        if data['Semanal'] == 'verde':
            cell_semanal.fill = verde
        else:
            cell_semanal.fill = rosa
            
        # Señal (aplicar color)
        cell_senal = ws.cell(row=row, column=6)
        if data['Señal'] == 'azul':
            cell_senal.fill = azul
        elif data['Señal'] == 'naranja':
            cell_senal.fill = naranja
    
    # Ajustar el ancho de las columnas
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
    
    # Agregar bordes a todas las celdas
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
    
    # Guardar el archivo en la ruta correcta
    wb.save(excel_path)
    
    # Abrir el archivo Excel automáticamente
    try:
        if os.name == 'nt':  # Windows
            os.startfile(excel_path)
        elif os.name == 'posix':  # macOS o Linux
            if os.uname().sysname == 'Darwin':  # macOS
                subprocess.run(['open', excel_path], capture_output=True, text=True)
            else:  # Linux
                subprocess.run(['xdg-open', excel_path], capture_output=True, text=True)
    except Exception:
        pass  # Silenciar cualquier error al abrir el archivo

def format_roc(roc_value):
    """Formatea el valor ROC con color según sea positivo o negativo"""
    if roc_value > 0:
        return f"{Colors.GREEN}{roc_value:.2f}%{Colors.ENDC}"
    else:
        return f"{Colors.RED}{roc_value:.2f}%{Colors.ENDC}"

def format_signal(signal):
    """Formatea la señal con color según su valor"""
    if signal == 'verde':
        return f"{Colors.GREEN}V{Colors.ENDC}"
    elif signal == 'rosa':
        return f"{Colors.RED}R{Colors.ENDC}"
    elif signal == 'amarillo':
        return f"{Colors.YELLOW}A{Colors.ENDC}"
    elif signal == 'azul':
        return f"{Colors.BLUE}A{Colors.ENDC}"
    elif signal == 'naranja':
        return f"{Colors.YELLOW}N{Colors.ENDC}"
    return signal

def main():
    """Función principal del programa"""
    print_header()
    results = []
    mensaje_sin_datos = 'Sin datos históricos suficientes'
    
    for i in range(0, len(TICKERS), BATCH_SIZE):
        batch = TICKERS[i:i + BATCH_SIZE]
        
        for ticker in batch:
            try:
                current_index = TICKERS.index(ticker) + 1
                weekly_data = download_data(ticker, '1wk')
                if not weekly_data.empty:
                    monthly_data = download_data(ticker, '1mo')
                    if monthly_data.empty:
                        results.append({
                            'Ticker': ticker,
                            'ROC': mensaje_sin_datos,
                            'Mensual': mensaje_sin_datos,
                            'Semanal': mensaje_sin_datos,
                            'Trimestral': mensaje_sin_datos,
                            'Señal': mensaje_sin_datos
                        })
                        if ticker == 'IBIT':
                            print(f"⚠️ No hay datos mensuales disponibles para {ticker}")
                        continue
                    weekly_data = calculate_indicators(weekly_data)
                    monthly_data = calculate_indicators(monthly_data)
                    monthly_data = calculate_trimestral_macd(monthly_data)
                    weekly_data = calculate_cross_macd(weekly_data)
                    roc_value = weekly_data['ROC'].iloc[-1]
                    macd_hist = monthly_data['MACD_Hist'].iloc[-1].item()
                    stoch_conditions = get_stoch_signal(monthly_data)
                    macd_weekly = get_macd_signal(weekly_data)
                    trimestral_signal = get_trimestral_signal(monthly_data)
                    cross_signal = get_cross_signal(weekly_data)
                    monthly_color = 'verde' if macd_hist > 0 else 'rosa'
                    weekly_color = 'verde' if macd_weekly == 'alza' else 'rosa'
                    
                    # Formatear la salida con colores
                    print(f"[{current_index}/{len(TICKERS)}] {Colors.BOLD}{ticker}{Colors.ENDC} | ROC: {format_roc(roc_value)} | T: {format_signal(trimestral_signal)} | M: {format_signal(monthly_color)} | S: {format_signal(weekly_color)} | Señal: {format_signal(cross_signal) if cross_signal else '-'}")
                    
                    results.append({
                        'Ticker': ticker,
                        'ROC': round(roc_value, 2) if not np.isnan(roc_value) else 0,
                        'Mensual': monthly_color,
                        'Semanal': weekly_color,
                        'Trimestral': trimestral_signal,
                        'Señal': cross_signal
                    })
                else:
                    results.append({
                        'Ticker': ticker,
                        'ROC': mensaje_sin_datos,
                        'Mensual': mensaje_sin_datos,
                        'Semanal': mensaje_sin_datos,
                        'Trimestral': mensaje_sin_datos,
                        'Señal': mensaje_sin_datos
                    })
                    if ticker == 'IBIT':
                        print(f"⚠️ No hay datos disponibles para {ticker}")
            except Exception as e:
                results.append({
                    'Ticker': ticker,
                    'ROC': mensaje_sin_datos,
                    'Mensual': mensaje_sin_datos,
                    'Semanal': mensaje_sin_datos,
                    'Trimestral': mensaje_sin_datos,
                    'Señal': mensaje_sin_datos
                })
                if ticker == 'IBIT':
                    print(f"❌ Error procesando {ticker}: Faltan datos históricos suficientes")
                continue

    if not results:
        print(f"{Colors.RED}No se pudieron procesar datos para ningún ticker{Colors.ENDC}")
        return

    df_results = pd.DataFrame(results)
    
    # Separar los resultados en numéricos y no numéricos
    df_numeric = df_results[df_results['ROC'].apply(lambda x: isinstance(x, (int, float)))]
    df_non_numeric = df_results[~df_results['ROC'].apply(lambda x: isinstance(x, (int, float)))]
    
    # Ordenar solo los resultados numéricos
    df_numeric = df_numeric.sort_values('ROC', ascending=False)
    
    # Concatenar los resultados ordenados
    df_results = pd.concat([df_numeric, df_non_numeric])
    
    # Cálculos adicionales solo con datos válidos
    valid_roc = df_numeric['ROC']
    roc_mean = valid_roc.mean() if not valid_roc.empty else 0
    roc_std = valid_roc.std() if not valid_roc.empty else 0
    positive_tickers = len(valid_roc[valid_roc > 0])
    positive_percentage = (positive_tickers / len(valid_roc)) * 100 if not valid_roc.empty else 0
    
    # Imprimir resumen mejorado con colores
    print(f"\n{Colors.BOLD}Resumen:{Colors.ENDC} Procesados: {len(valid_roc)}/{len(TICKERS)} | Mayor ROC: {format_roc(valid_roc.max()) if not valid_roc.empty else 'N/A'} | Menor ROC: {format_roc(valid_roc.min()) if not valid_roc.empty else 'N/A'}")
    print(f"{Colors.BOLD}Estadísticas:{Colors.ENDC} Media ROC: {format_roc(roc_mean)} | Volatilidad: {Colors.YELLOW}{roc_std:.2f}%{Colors.ENDC} | Positivos: {Colors.GREEN}{positive_percentage:.1f}%{Colors.ENDC} ({positive_tickers}/{len(valid_roc)})")
    
    try:
        export_to_excel(df_results)
    except Exception as e:
        print(f"{Colors.RED}Error al exportar a Excel: {str(e)}{Colors.ENDC}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n{Colors.YELLOW}Proceso interrumpido por el usuario{Colors.ENDC}")
    except Exception as e:
        print(f"\n{Colors.RED}Error inesperado: {str(e)}{Colors.ENDC}")
    finally:
        print(f"\n{Colors.BLUE}Fin del programa{Colors.ENDC}") 